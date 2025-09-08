import pandas as pd
from playwright.sync_api import sync_playwright
import time, random
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def fetch_all_categories(word, page):
    url = f"https://woordenlijst.org/zoeken/?q={word}"
    try:
        page.goto(url)
        page.wait_for_selector("#lemma_region", timeout=5000)
    except:
        return []

    results = []
    regions = page.query_selector_all("#lemma_region")

    for region in regions:
        categorie_elem = region.query_selector(".lemma_head_cat")
        categorie = categorie_elem.inner_text().strip() if categorie_elem else "not found"

        name_elem = region.query_selector('span[itemprop="name"]')
        name_found = name_elem.inner_text().strip() if name_elem else "missing"

        article = ""
        if "zelfstandig naamwoord" in categorie.lower():
            spans = region.query_selector_all("span")
            article = spans[0].inner_text().strip() if spans else ""

        results.append((article, name_found, categorie))

    return results


def filter_results(word, site_results):
    match = None
    for r in site_results:
        if r[1].lower() == word.lower():
            match = r
            break

    if not match:
        return [("", word, "not found")]

    base_cat = match[2].lower()
    noun = next((r for r in site_results if "zelfstandig naamwoord" in r[2].lower()), None)
    verb = next((r for r in site_results if "werkwoord" in r[2].lower()), None)

    if "zelfstandig naamwoord" in base_cat and verb:
        return [match, verb]
    elif "werkwoord" in base_cat and noun:
        return [match, noun]
    elif ("werkwoord" not in base_cat and "zelfstandig naamwoord" not in base_cat) and noun:
        return [match, noun]
    else:
        return [match]


def build_vocabulary(input_file="vocabulary_new.xlsx", output_file="vocabulary_mined.xlsx"):
    df = pd.read_excel(input_file)
    all_results = []
    found_flags = []

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page()

        for word in df["Nederlands"].tolist():
            site_results = fetch_all_categories(word, page)

            if site_results:
                final_row = filter_results(word, site_results)
                found_flags.append("Yes")
            else:
                final_row = [("", word, "not found")]
                found_flags.append("No")

            all_results.append(final_row)
            time.sleep(random.uniform(0.5, 1.5))

        browser.close()

    max_len = max(len(r) for r in all_results)

    columns = {}
    for i in range(max_len):
        columns[f"Article_{i+1}"] = []
        columns[f"Name_{i+1}"] = []
        columns[f"Category_{i+1}"] = []

    for row in all_results:
        for i in range(max_len):
            if i < len(row):
                a, n, c = row[i]
            else:
                a, n, c = "", "", ""
            columns[f"Article_{i+1}"].append(a)
            columns[f"Name_{i+1}"].append(n)
            columns[f"Category_{i+1}"].append(c)

    for col, values in columns.items():
        df[col] = values

    df["Found_in_site"] = found_flags

    # Sauvegarde Excel
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        df.to_excel(writer, index=False)
        

    # Charger et styliser
    wb = load_workbook(output_file)
    ws = wb.active

    yellow = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")

    # Identifier colonnes par nom
    header_row = [cell.value for cell in ws[1]]
    target_cols = ["Article_1", "Name_1", "Article_2", "Name_2"]

    for col_name in target_cols:
        if col_name in header_row:
            col_idx = header_row.index(col_name) + 1
            for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2, max_row=ws.max_row):
                for c in cell:
                    c.fill = yellow

    wb.save(output_file)
    print(f"✅ Filtered & styled Excel saved as {output_file}")


if __name__ == "__main__":
    build_vocabulary()
